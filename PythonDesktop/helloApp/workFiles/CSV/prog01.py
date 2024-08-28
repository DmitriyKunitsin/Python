import openpyxl

name_file ='text.xlsx'

wb = openpyxl.load_workbook(name_file)
print(wb.sheetnames)
print(type(wb))

currSheet = wb['Лист1']
print(currSheet)
print(type(currSheet))
print('\n')
currSheet = wb[wb.sheetnames[0]]
print(currSheet)
print(type(currSheet))
print(currSheet.title)
## Печатаем нужную ячейку
print('\n')
var1 = currSheet['A1']
print(var1.value)
print('\n')
print(currSheet['B1'].value)
print('\n')
var2 = currSheet.cell(row=2, column=2)
print(var2.value)
## Максимальное колличество строк и столбцов
print('\n')
print(currSheet.max_row)
print(currSheet.max_column)

for i in range(currSheet.max_row):
    print('---Новая строка---')
    for j in range(currSheet.max_column):
        var = currSheet.cell(row=i+1,column=j+1)
        print(var.value)
    print('---Конец строки---')